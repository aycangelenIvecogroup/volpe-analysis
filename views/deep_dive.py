import streamlit as st
import pandas as pd

from ui.helpers import h
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule

# -----------------------------
# LOAD DATA
# -----------------------------
@st.cache_data
def load_data():
    return pd.read_excel("customer_amount_layer_clean.xlsx")


# -----------------------------
# PAGE RENDER
# -----------------------------
def render_deep_dive(controls):

    st.title("🔍 Deep Dive")
    st.caption("Select customers and metrics to explore performance indicators.")
    st.divider()

    df = load_data()

    # -----------------------------
    # CONTROLS
    # -----------------------------
    month = controls["month"]
    scenario = controls["scenario"]

    # Scenario‑aware defaults
    if scenario == "ACT":
        default_metrics = [
            f"{month}_AGM%",
            f"{month}_Coverage_vs_B26%"
        ]
    elif scenario == "B26":
        default_metrics = ["B26_AGM%"]
    else:
        default_metrics = [
            f"{scenario}_AGM%",
            f"{scenario}_Coverage_vs_B26%"
        ]

    # -----------------------------
    # CUSTOMER SELECTION
    # -----------------------------
    customers = st.multiselect(
        "Select Customers",
        sorted(df["CUSTOMER MERGE"].unique()),
        help=h("customer")
    )

    if not customers:
        st.info("Please select at least one customer.")
        st.stop()

    df = df[df["CUSTOMER MERGE"].isin(customers)]
    st.divider()

    # -----------------------------
    # METRIC SELECTION
    # -----------------------------
    numeric_cols = [
        c for c in df.columns
        if df[c].dtype != "object" and c != "CUSTOMER MERGE"
    ]

    metrics = st.multiselect(
        "Select Metrics",
        numeric_cols,
        default=[c for c in default_metrics if c in numeric_cols],
    )

    if not metrics:
        st.warning("Please select at least one metric.")
        st.stop()

    st.divider()

    # -----------------------------
    # DATA PREPARATION
    # -----------------------------
    view_df = df[["CUSTOMER MERGE"] + metrics].copy()

    # Numeric cleanup (THIS fixes ugly numbers)
    for col in metrics:
        if col.endswith("%") or "GAP" in col or "YoY" in col:
            view_df[col] = pd.to_numeric(view_df[col], errors="coerce").round(1)
        elif col.endswith("_TN") or col.endswith("_Units"):
            view_df[col] = pd.to_numeric(view_df[col], errors="coerce").round(1)

    # -----------------------------
    # RESULT TABLE
    # -----------------------------
    st.subheader("📊 Deep Dive Results")

    column_config = {
        "CUSTOMER MERGE": st.column_config.TextColumn("Customer")
    }

    for col in metrics:
        if col.endswith("_AGM%"):
            column_config[col] = st.column_config.NumberColumn(
                "AGM %",
                format="%.1f %"
            )
        elif col.endswith("_SGM%"):
            column_config[col] = st.column_config.NumberColumn(
                "SGM %",
                format="%.1f %"
            )
        elif col.endswith("_Coverage_vs_B26%"):
            column_config[col] = st.column_config.NumberColumn(
                "Coverage %",
                format="%.1f %"
            )
        elif col.endswith("_Units"):
            column_config[col] = st.column_config.NumberColumn(
                "Units",
                format="%.1f"
            )
        elif col.endswith("_TN"):
            column_config[col] = st.column_config.NumberColumn(
                "Net Sales (TN)",
                format="%.1f"
            )
        elif "GAP" in col:
            column_config[col] = st.column_config.NumberColumn(
                col,
                format="%.1f pp"
            )
        elif "YoY" in col:
            column_config[col] = st.column_config.NumberColumn(
                col,
                format="%.1f %"
            )
        else:
            column_config[col] = st.column_config.NumberColumn(col)

    st.dataframe(
        view_df,
        use_container_width=True,
        height=520,
        column_config=column_config
    )

    # -----------------------------
    # EXCEL EXPORT
    # -----------------------------
    st.divider()

    def to_excel(df):
        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Deep Dive")

            wb = writer.book
            ws = wb["Deep Dive"]
            ws.freeze_panes = "B2"

            for cell in ws[1]:
                cell.font = Font(bold=True)

            ws.auto_filter.ref = ws.dimensions

            color_scale = ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B"
            )

            ws.conditional_formatting.add(
                f"B2:{ws.cell(ws.max_row, ws.max_column).coordinate}",
                color_scale
            )

        return output.getvalue()

    st.download_button(
        label="📥 Download Deep Dive as Excel",
        data=to_excel(view_df),
        file_name=f"deep_dive_{scenario}_{month}_{len(customers)}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
import pandas as pd

